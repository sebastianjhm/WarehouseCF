from flask import Flask, request, send_file, jsonify
from flask_cors import CORS, cross_origin
import openpyxl
from openpyxl.utils import get_column_letter
import xlsxwriter
import pyomo.environ as pyo
from pyomo.opt import SolverFactory
import copy
import pyutilib.subprocess.GlobalData
from Rutas import Rutas
from Racks import Racks

## Librería para que pyomo corra dentro de Flask
pyutilib.subprocess.GlobalData.DEFINE_SIGNAL_HANDLERS_DEFAULT = False

app = Flask(__name__)
cors = CORS(app)

@app.route('/postFileAllocation', methods=['POST'])
@cross_origin()
def post_allocation():
    
    ## SAVE EXCEL FILE THAT ARRIVE OF THE FRONT
    receivedFile = request.files["myExcelFileAlloc"]
    
    ## READ DATA OPENPYXL
    set_REF , param_AnchoCaja, param_TipoDist, param_Espacios, param_Demanda, param_Frecuencia, set_RACKS, param_TipoRack, param_Pared, param_Utilizacion, param_Costo = read_data_XLSX_Alloc( receivedFile )
    
    ## DECLARE SOLVER, CREATE A PYOMO ABSTRACT MODEL, DECLARE LINEAR MODEL
    opt = SolverFactory('cplex', executable="C:\\Program Files\\IBM\\ILOG\\CPLEX_Studio128\\cplex\\bin\\x64_win64\\cplex")
    model = pyo.AbstractModel()
    create_lineal_model_Alloc( model, set_REF , param_AnchoCaja, param_TipoDist, param_Espacios, param_Demanda, param_Frecuencia, set_RACKS, param_TipoRack, param_Pared, param_Utilizacion, param_Costo )
    
    ## CREATE AN INSTANCE OF THE MODEL, SOLVE PRINT RESULTS BY CONSOLE
    instance = model.create_instance()
    opt.options['timelimit'] = 20
    results = opt.solve(instance, tee = False)
    #instance.display()
    
    ## SAVE JSON ROUTES
    Racks["fo"] = 0
    Rutas["racks"] = []
    save_racks_Alloc( instance )
    
    ## PRINT RESULTS IN CONSOLE
    print_results_in_console_Alloc( instance )
    
    ## CREATE EXCEL FILE AND PRINT RESULTS WITH 'XLSXWRITER'
    workbook = xlsxwriter.Workbook('Results_Allocation.xlsx')
    print_results_XLSX_Alloc( workbook, instance )
     
    return send_file('Results_Allocation.xlsx', as_attachment=True)
#fed
    
@app.route('/racks', methods = ['GET'])
@cross_origin()
def get_racks_Alloc():
    return jsonify(Racks)
#fed




def save_racks_Alloc( instance ):
    
    Racks["fo"] = pyo.value(instance.FO)
    new_rack = {"id":0, "hileras": False, "referencias": []}
    
    for rack in instance.RACKS:
        nr = copy.deepcopy( new_rack )
        nr["id"] = rack
        if (instance.y[rack] == 1):
            nr["hileras"] = True
        else:
            nr["hileras"] = False
        #fi
        for ref in instance.REF:
            if ( instance.x[ref,rack] == 1 ):
                nr["referencias"].append(ref)
            #fi
        #rof
        Racks["racks"].append(nr)
    #rof
    print(Racks)
#fed
    
def print_results_in_console_Alloc( instance ):
    print("\n\n")
    print("SOLUCIÓN DEL EJERCICIO")
    print("--------------------------")
    print("\n")
    print("Función Objetivo: ",pyo.value(instance.FO))
    print("\n")
    for ref in instance.REF:
        for rack in instance.RACKS:
            if ( instance.x[ref,rack] == 1 ):
                print("La referencia ", ref ," en el rack", rack)
            #fi
        #rof
    #rof
#fed

def read_data_XLSX_Alloc( receivedFile ):
    ## --------------------- READ DATA EXCEL WHIT OPENPYXL -----------------------------------
    
    ## OPEN FILE AND SAVE IN VARIABLE (archivo)
    archivo = openpyxl.load_workbook( receivedFile, data_only = True)

    ## READ SHEETS OF FILE AND SELECT THE THE FIRST SHEET
    sheets = archivo.sheetnames
    sheetDatos = archivo[sheets[0]]
    
    ## REFERENCES (set_REF)
    set_REF = []
    column = sheetDatos['B']
    for x in range(len(column)): 
        if (type(column[x].value) != str and column[x].value != None):
            set_REF.append(column[x].value)
        #fi
    #rof 
    
    ## BOXES WIDTH (param_AnchoCaja)
    lect = []
    column = sheetDatos['C']
    for x in range(len(column)): 
        if (type(column[x].value) != str and column[x].value != None):
            lect.append(column[x].value)
        #fi
    #rof
    param_AnchoCaja = dict(zip(set_REF, lect))
    
    ## IF REFERENCE IS IN ROWS (param_TipoDist)
    lect = []
    column = sheetDatos['D']
    for x in range(len(column)): 
        if (type(column[x].value) != str and column[x].value != None):
            lect.append(column[x].value)
        #fi
    #rof
    param_TipoDist = dict(zip(set_REF, lect))
    
    ## NUMBERS SPACES OF REFERENCE (param_Espacios)
    lect = []
    column = sheetDatos['E']
    for x in range(len(column)): 
        if (type(column[x].value) != str and column[x].value != None):
            lect.append(column[x].value)
        #fi
    #rof
    param_Espacios = dict(zip(set_REF, lect))
    
    ## DEMAND REFERENCE (param_Demanda)
    lect = []
    column = sheetDatos['F']
    for x in range(len(column)): 
        if (type(column[x].value) != str and column[x].value != None):
            lect.append(column[x].value)
        #fi
    #rof
    param_Demanda= dict(zip(set_REF, lect))
    
    ## FREQUENCY REFERENCE (param_Frecuencia)
    lect = []
    column = sheetDatos['G']
    for x in range(len(column)): 
        if (type(column[x].value) != str and column[x].value != None):
            lect.append(column[x].value)
        #fi
    #rof
    param_Frecuencia = dict(zip(set_REF, lect))
    
    ## =======================================================================
    ## RACKS (set_RACKS)
    set_RACKS = []
    column = sheetDatos['L']
    for x in range(len(column)): 
        if (type(column[x].value) != str and column[x].value != None):
            set_RACKS.append(column[x].value)
        #fi
    #rof
    
    ## IF THE RACK HAVE BACKGROUND (param_TipoRack)
    lect = []
    column = sheetDatos['M']
    for x in range(len(column)): 
        if (type(column[x].value) != str and column[x].value != None):
            lect.append(column[x].value)
        #fi
    #rof
    param_TipoRack = dict(zip(set_RACKS, lect))
    
    ## IF THE RACK IS IN WALL (param_Pared)
    lect = []
    column = sheetDatos['N']
    for x in range(len(column)): 
        if (type(column[x].value) != str and column[x].value != None):
            lect.append(column[x].value)
        #fi
    #rof
    param_Pared = dict(zip(set_RACKS, lect))
    
    ## IF THE RACK IS IN WALL (param_Utilizacion)
    lect = []
    column = sheetDatos['O']
    for x in range(len(column)): 
        if (type(column[x].value) != str and column[x].value != None):
            lect.append(column[x].value)
        #fi
    #rof
    param_Utilizacion = dict(zip(set_RACKS, lect))
    
    ## COST OF ALLOCATE IN THE RACK (param_Costo)
    lect = []
    column = sheetDatos['P']
    for x in range(len(column)): 
        if (type(column[x].value) != str and column[x].value != None):
            lect.append(column[x].value)
        #fi
    #rof
    param_Costo = dict(zip(set_RACKS, lect))
    
    
    return( set_REF , param_AnchoCaja, param_TipoDist, param_Espacios, param_Demanda, param_Frecuencia, set_RACKS, param_TipoRack, param_Pared, param_Utilizacion, param_Costo )
#fed

def create_lineal_model_Alloc( model, set_REF , param_AnchoCaja, param_TipoDist, param_Espacios, param_Demanda, param_Frecuencia, set_RACKS, param_TipoRack, param_Pared, param_Utilizacion, param_Costo ):
    ## --------------------- CONJUNTOS ----------------------------
    model.REF = pyo.Set( initialize = set_REF )
    model.RACKS = pyo.Set( initialize = set_RACKS )
    
    ## ---------------------- PARÁMETROS ----------------------------
    model.AnchoCaja = pyo.Param( model.REF, initialize = param_AnchoCaja )
    model.TipoDist = pyo.Param( model.REF, initialize = param_TipoDist )
    model.Espacios = pyo.Param( model.REF, mutable=True, initialize = param_Espacios )
    model.TipoRack = pyo.Param( model.RACKS, initialize = param_TipoRack )
    model.Pared = pyo.Param( model.RACKS, initialize = param_Pared)
    model.Utilizacion = pyo.Param( model.RACKS, initialize = param_Utilizacion )
    
    model.Demanda  = pyo.Param( model.REF, initialize = param_Demanda )
    model.Frecuencia  = pyo.Param( model.REF, initialize = param_Frecuencia )
    model.Costo = pyo.Param( model.RACKS, initialize = param_Costo )
    
    ## ---------------------- VARIABLES ----------------------------
    model.x = pyo.Var( model.REF, model.RACKS, domain = pyo.Binary )
    model.y = pyo.Var( model.RACKS, domain = pyo.Binary )
    
    ## ---------------------- FUNCIÓN OBJETIVO ----------------------------
    def ObjFunc( model ):
        return sum(model.Demanda[ref]*model.Frecuencia[ref]*model.Costo[rack]*model.x[ref,rack] for ref in model.REF for rack in model.RACKS )
    #fed
    model.FO = pyo.Objective( rule = ObjFunc )
    
    ## ---------------------- RESTRICCIONES ----------------------------
    def r1( model, ref ):
        return sum( model.x[ref,rack] for rack in model.RACKS ) == 1
    #fed
    model.r1 = pyo.Constraint( model.REF, rule = r1 )
    
    def r2( model, rack ):
        return sum( model.AnchoCaja[ref]*model.x[ref,rack] for ref in model.REF ) <= 250*(1-model.y[rack]) + 750*(model.y[rack])
    #fed
    model.r2 = pyo.Constraint( model.RACKS, rule = r2)
    
    def r3( model, ref, rack):
        return model.x[ref,rack] <= (model.y[rack]*model.TipoDist[ref]) + ((1-model.y[rack])*(1-model.TipoDist[ref]))
    #fed
    model.r3 = pyo.Constraint( model.REF, model.RACKS, rule = r3 )
    
    def r6( model, ref, rack ):
        return model.Espacios[ref]*model.x[ref,rack] <= model.TipoRack[rack]
    #fed
    model.r6 = pyo.Constraint( model.REF, model.RACKS, rule = r6 )
    
    def r8( model, rack ):
        return model.y[rack] <= 1 - model.Pared[rack]
    #def
    model.r8  = pyo.Constraint( model.RACKS, rule = r8 )
     
    def r9( model, ref, rack ):
        return model.x[ref,rack] <= model.Utilizacion[rack]
    #def
    model.r9 = pyo.Constraint( model.REF, model.RACKS, rule = r9 )
#fed


def print_results_XLSX_Alloc( workbook, instance ):
    ## -------------------------- SAVE RESULTS IN EXCEL FILE -------------------------------------

    ## STYLES
    merge_format = workbook.add_format({
        'bold': 1,
        'align': 'center',
        'valign': 'vcenter'})
    cell_format = workbook.add_format({'align': 'center'})
        
    ## CREATE SHEET
    worksheet = workbook.add_worksheet('Res. REFERENCES')
    worksheet.set_column(0, 0, 16)
    worksheet.set_column(1, 1, 11) ## Width: set_column(first_col, last_col, width, cell_format, options)
    worksheet.set_column(2, 2, 10)
    worksheet.set_column(3, 3, 9.5)
    
    ## PRINT RESULTS
    worksheet.merge_range(0, 0, 0, 3, "SOLUCIÓN ASIGNACIÓN REFERENCIA", merge_format)
    worksheet.merge_range(1, 0, 1, 3, "", merge_format)
    worksheet.write(2, 0, "Función Objetivo: ", merge_format)
    worksheet.merge_range(2, 1, 2, 3, pyo.value(instance.FO), cell_format)
    worksheet.merge_range(3, 0, 3, 3, "", merge_format)
    worksheet.merge_range(4, 0, 4, 3, "Asignación", merge_format)
    
    row = 5
    for ref in instance.REF:
        for rack in instance.RACKS:
            if ( instance.x[ref,rack] == 1 ):
                worksheet.write(row, 0, "La referencia", cell_format)
                worksheet.write(row, 1, ref, cell_format)
                worksheet.write(row, 2, "en el rack", cell_format)
                worksheet.write(row, 3, rack, cell_format)
            #fi
        #rof
        row = row + 1
    #rof
    
    
    
    ## CREATE SHEET 2
    worksheet2 = workbook.add_worksheet('Res. RACKS')
    worksheet2.set_column(0, 2, 16) ## Width: set_column(first_col, last_col, width, cell_format, options)
    
    
    ## PRINT RESULTS
    worksheet2.merge_range(0, 0, 0, 2, "SOLUCIÓN ASIGNACIÓN RACKS", merge_format)
    worksheet2.write(1, 0, "Función Objetivo: ", merge_format)
    worksheet2.merge_range(1, 1, 1, 2, pyo.value(instance.FO), cell_format)
    worksheet2.merge_range(2, 0, 2, 2, "", merge_format)
    worksheet2.merge_range(3, 0, 3, 2, "Asignación", merge_format)
    worksheet2.write(4, 0, "RACK", merge_format)
    worksheet2.write(4, 1, "REFERENCIA", merge_format)
    worksheet2.write(4, 2, "¿HILERAS?", merge_format)
    
    row = 6
    for rack in instance.RACKS:
        ind = False
        for ref in instance.REF:
            if ( instance.x[ref,rack] == 1 ):
                worksheet2.write(row, 0, rack, cell_format)
                worksheet2.write(row, 1, ref, cell_format)
                if ( instance.y[rack] == 1 ):
                    worksheet2.write(row, 2, "si", cell_format)
                else:
                    worksheet2.write(row, 2, "no", cell_format)
                #fi
                row = row + 1
                ind = True
            #fi
        #rof
        if ind == True:
            row = row + 1
        #fi
    #rof
    
    ## CLOSE THE BOOK
    workbook.close()
#fed




































@app.route('/postFilePicking', methods=['POST'])
@cross_origin()
def post_picking():
    
    ## SAVE EXCEL FILE THAT ARRIVE OF THE FRONT
    receivedFile = request.files["myExcelFile"]
    
    
    ## READ DATA OPENPYXL
    set_Nodos, set_Ords, set_Referencias, param_NOD_REF, set_Ordenes, set_R, param_Distancia = read_data_XLSX_Picking( receivedFile )
    
    
    ## DECLARE SOLVER, CREATE A PYOMO ABSTRACT MODEL, DECLARE LINEAR MODEL
    opt = SolverFactory('cplex', executable="C:\\Program Files\\IBM\\ILOG\\CPLEX_Studio128\\cplex\\bin\\x64_win64\\cplex")
    #opt = SolverFactory('glpk')
    model = pyo.AbstractModel()
    create_lineal_model_Picking( model, set_Nodos, set_Ords, set_Referencias, param_NOD_REF, set_Ordenes, set_R, param_Distancia )
    
    
    ## CREATE AN INSTANCE OF THE MODEL, SOLVE PRINT RESULTS BY CONSOLE
    instance = model.create_instance()
    opt.options['timelimit'] = 10
    results = opt.solve(instance, tee=False)
    #instance.display()
    
    ## PRINT RESULTS IN CONSOLE
    print_results_in_console_Picking( instance )
    
    ## SAVE JSON ROUTES
    Rutas["distanciaTotal"] = 0
    Rutas["ruta"] = []
    save_routes_Picking( instance )
    
    
    ## CREATE EXCEL FILE AND PRINT RESULTS WITH 'XLSXWRITER'
    workbook = xlsxwriter.Workbook('Results_Picking.xlsx')
    print_results_XLSX_Picking( workbook, instance )
    
    
    return send_file('Results_Picking.xlsx', as_attachment=True)

#fed
    




@app.route('/rutas', methods = ['GET'])
@cross_origin()
def get_routes_picking():
    return jsonify(Rutas)
#fed
    





################################################################################################################
def print_results_in_console_Picking( instance ):
    
    print("\n\n")
    print("SOLUCIÓN DEL EJERCICIO")
    print("--------------------------")
    print("\n")
    print("Distancia Total: ",pyo.value(instance.FO))
    print("\n")
    for o in instance.O:
        print("--------- Orden ",o,"----------")
        print("Distancia Recorrida: ",instance.Dist_ORD[o].value)
        print("Ruta: ")
        for i in instance.ORDENES[o]:
            for j in instance.ORDENES[o]:
                if ( instance.x[o,i,j].value == 1):
                    print("Del nodo ",i," al nodo",j)
                #fi
            #rof
        #rof
    #rof
#fed
    
def save_routes_Picking( instance ):
    
    nueva_ruta = {"id":0, "distanciaRuta": 0, "recorrido": []}
    Rutas["distanciaTotal"] = pyo.value(instance.FO)
    
    for o in instance.O:
        nr = copy.deepcopy( nueva_ruta )
        nr["id"] = o
        nr["distanciaRuta"] = instance.Dist_ORD[o].value
        for i in instance.ORDENES[o]:
            for j in instance.ORDENES[o]:
                if ( instance.x[o,i,j].value == 1):
                    r = [i, j]
                    nr["recorrido"].append(r)
                #fi
            #rof
        #rof
        Rutas["ruta"].append(nr)
    #rof
    print(Rutas)
    
#fed
################################################################################################################   
    
    
    
######################################################################################################################################  
def read_data_XLSX_Picking( receivedFile ):
    
    ## --------------------- READ DATA EXCEL WHIT OPENPYXL -----------------------------------
    
    ## OPEN FILE AND SAVE IN VARIABLE (archivo)
    archivo = openpyxl.load_workbook(receivedFile, data_only = True)

    ## READ SHEETS OF FILE AND SELECT THE THE FIRST SHEET
    sheets = archivo.sheetnames
    sheetDatos = archivo[sheets[0]]

    ## NODES (set_Nodos)
    set_Nodos = []
    column = sheetDatos['A']
    for x in range(len(column)): 
        if (type(column[x].value) != str and column[x].value != None):
            set_Nodos.append(column[x].value)
        #fi
    #rof

    ## ORDERS (set_Ords)
    set_Ords = []
    column = sheetDatos['B']
    for x in range(len(column)): 
        if (type(column[x].value) != str and column[x].value != None):
            set_Ords.append(column[x].value)
        #fi
    #rof


    ## REFERENCES (set_Referencias)
    set_Referencias = []
    column = sheetDatos['C']
    for x in range(len(column)): 
        if (type(column[x].value) != str and column[x].value != None):
            set_Referencias.append(column[x].value)
        #fi
    #rof

    ## LOCATION OF EACH REFERENCE (param_NOD_REF)
    lect = []
    column = sheetDatos['F']
    for x in range(len(column)): 
        if (type(column[x].value) != str and column[x].value != None):
            lect.append(column[x].value)
        #fi
    #rof
    param_NOD_REF = dict(zip(set_Referencias, lect))


    ## Ordenes y Variable auxiliar (set_Ordenes y set_R)
    lect = []
    for i in range(len(set_Ords)):
        column = sheetDatos[get_column_letter( 8 + i )]
        orden = []
        for x in range(len(column)): 
            if (type(column[x].value) != str and column[x].value != None):
                orden.append(column[x].value)
            #fi
        #rof
        lect.append(orden)
    #rof
    set_Ordenes = dict(zip(set_Ords, lect))

    lectR = copy.deepcopy(lect) ## Esto es demencial
    for x in lectR:
        x.remove(0)
    #rof
    set_R = dict(zip(set_Ords, lectR))


    ## DISTANCE MATRIX
    param_Distancia = []
    sheetDistancias = archivo[sheets[1]]
    for i in  range(sheetDistancias.min_row + 1,sheetDistancias.max_row + 1):
        y = []
        for j in range(sheetDistancias.min_column + 1, sheetDistancias.max_column + 1):
            y.append(sheetDistancias.cell(row = i, column = j).value)
        #rof
        param_Distancia.append(y)
    #rof
    ## -------------------------------------------------------------------------------
    
    return( set_Nodos, set_Ords, set_Referencias, param_NOD_REF, set_Ordenes, set_R, param_Distancia )
#fed
######################################################################################################################################    
    

######################################################################################################################################
def create_lineal_model_Picking( model, set_Nodos, set_Ords, set_Referencias, param_NOD_REF, set_Ordenes, set_R, param_Distancia ):
    

    ## ------------------- SETS --------------------------------
    model.NODOS = pyo.Set( initialize = set_Nodos )
    model.O = pyo.Set( initialize = set_Ords )
    model.REF = pyo.Set( initialize = set_Referencias )
    def setORDENES(model, i):
        return(list(set_Ordenes[i]))
    #fed
    model.ORDENES = pyo.Set(model.O, initialize = setORDENES)
    def setR(model, i):
        return(list(set_R[i]))
    #fed
    model.R = pyo.Set(model.O, initialize = set_R)



    def junte(model):
        return((o,i,j) for o in model.O for i in model.ORDENES[o] for j in model.ORDENES[o] )
    #fed
    model.OROR = pyo.Set(dimen =3, initialize = junte )

    def junte2(model):
        return((o,i) for o in model.O for i in model.ORDENES[o] )
    #fed
    model.OX = pyo.Set(dimen = 2, initialize = junte2 )

    def junte3(model):
        return((o,i,j) for o in model.O for i in model.R[o] for j in model.R[o] )
    #fed
    model.ORR = pyo.Set(dimen = 3, initialize = junte3 )

    def junte4(model):
        return((o,i) for o in model.O for i in model.R[o] )
    #fed
    model.OR  = pyo.Set(dimen = 2, initialize = junte4 )


    ## --------------------- PARAMETERS -------------------------------------
    def paramDistancia(model, i, j):
        return(param_Distancia[i][j])
    #fed
    model.Distancia = pyo.Param(model.NODOS,model.NODOS, initialize = paramDistancia)
    
    def paramNodRef(model, i):
        return(param_NOD_REF[i])
    #fed
    model.Nod_Ref = pyo.Param(model.REF, initialize = paramNodRef) 

    def paramDistR(model, i ,j):
        return model.Distancia[model.Nod_Ref[i],model.Nod_Ref[j]]
    #fed
    model.Distancia_R = pyo.Param(model.REF, model.REF, initialize = paramDistR, mutable = True)

    ## ----------------------- VARIABLES ---------------------------------------
    model.x = pyo.Var(model.OROR, domain = pyo.Binary)
    model.aux = pyo.Var(model.OR)
    model.Dist_ORD = pyo.Var(model.O)


    ## -------------------------- OBJECTIVE FUNCTION ------------------------------------
    def ObjFunc(model):
        return sum(model.Distancia_R[i,j]*model.x[o,i,j] for o in model.O for i in model.ORDENES[o] for j in model.ORDENES[o])
    #fed
    model.FO = pyo.Objective(rule = ObjFunc)


    ## --------------------------- RESTRICTIONS ----------------------------------------------
    def r1(model, o, i):
        return sum( model.x[o,i,j] for j in model.ORDENES[o]) == 1
    #fed
    model.r1 = pyo.Constraint( model.OX, rule = r1 )

    def r2(model, o, j):
        return sum( model.x[o,i,j] for i in model.ORDENES[o]) == 1
    #fed
    model.r2 = pyo.Constraint( model.OX,  rule = r2 )

    def r3(model, o, i):
        return model.x[o,i,i] == 0
    #fed
    model.r3 = pyo.Constraint( model.OX,  rule = r3 )

    def r4(model, o, i, j):
        if ( i != j ):
            return model.aux[o,i] - model.aux[o,j] + len(model.R[o])*model.x[o,i,j] <= len(model.R[o]) - 1 
        return pyo.Constraint.Skip
    #fed
    model.r4 = pyo.Constraint( model.ORR,  rule = r4 )

    def r5(model, o):
        return model.Dist_ORD[o] == sum( model.x[o,i,j]*model.Distancia_R[i,j] for i in model.ORDENES[o] for j in model.ORDENES[o] )
    #fed
    model.r5 = pyo.Constraint( model.O, rule = r5)
#fed
#########################################################################################################################################
    
    
    
    
    
#########################################################################################################################################
def print_results_XLSX_Picking( workbook, instance ):
    ## -------------------------- SAVE RESULTS IN EXCEL FILE -------------------------------------

    ## STYLES
    merge_format = workbook.add_format({
        'bold': 1,
        'align': 'center',
        'valign': 'vcenter'})
    cell_format = workbook.add_format({'align': 'center'})
        
    ## CREATE SHEET
    worksheet = workbook.add_worksheet('Resultados')
    worksheet.set_column(0, 0, 20) ## Width: set_column(first_col, last_col, width, cell_format, options)

    ## PRINT RESULTS
    worksheet.merge_range(0, 0, 0, 3, "SOLUCIÓN DEL EJERCICIO", merge_format)
    worksheet.write(1, 0, "Distancia Total: ", merge_format)
    worksheet.write(1, 1, pyo.value(instance.FO), cell_format)
    
    row=3
    col=0
    for o in instance.O:
        worksheet.merge_range(row, col, row, 3, "Ruta "+ str(o), merge_format)
        row += 1

        worksheet.write(row, col,"Distancia Recorrida: ", cell_format)
        worksheet.write(row, col + 1,instance.Dist_ORD[o].value, cell_format)
        row += 1
        
        worksheet.write(row, col,"Recorrido: ", cell_format)
        row += 1
        
        for i in instance.ORDENES[o]:
            for j in instance.ORDENES[o]:
                if ( instance.x[o,i,j].value == 1):
                    worksheet.write(row, col + 0,"Del rack ", cell_format)
                    worksheet.write(row, col + 1, i, cell_format)
                    worksheet.write(row, col + 2," al rack", cell_format)
                    worksheet.write(row, col + 3, j, cell_format)
                #fi
            #rof
            row += 1
        #rof
        row += 1
    #rof
        
        
    ## CLOSE THE BOOK
    workbook.close()
#fed
##########################################################################################################################################
    
    
    
if __name__ == '__main__':
    app.run(debug = False, port=5000)
#fi
