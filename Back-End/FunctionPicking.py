import openpyxl
from openpyxl.utils import get_column_letter
import xlsxwriter
import pyomo.environ as pyo
import copy
from Rutas import Rutas


################################################################################################################
def print_results_in_console_Picking( instance ):
    
    ## Guardar recorridos
    results = []
    for o in instance.O:
        route = []
        for i in instance.ORDENES[o]:
            for j in instance.ORDENES[o]:
                if ( instance.x[o,i,j].value == 1 ):
                    route.append([i, j])
                #fi
            #rof
        #rof
        results.append(route)
    #rof
    print(results)
    
    ## Ordenar Recorridos
    for i in range(len(results)):
        result = [results[i][0]]
        for _ in range(len(results[i])-1):
            aux = result[len(result)-1]
            for r in results[i]:
                if ( r[0] == aux[1] ):
                    result.append(r)
                    break
                #fi
            #rof
        #rof
        results[i] = result
    #rof
    
    ## Guardar rutas continuas
    cont_route = []
    for r in results:
        f = []
        for x in r:
            f.append(instance.Nod_Ref[x[0]])
        #rof
        f.append(instance.Nod_Ref[0])
        cont_route.append(f)
    #rof
    
    ## Imprimir
    print("\n\n")
    print("SOLUCIÓN DEL EJERCICIO(RACKS a RACKS)")
    print("--------------------------")
    print("\n")
    print("Distancia Total: ", pyo.value(instance.FO))
    print("\n")
    for o in instance.O:
        print("--------- Orden ",o,"----------")
        print("Distancia Recorrida: ",instance.Dist_ORD[o].value)
        print("Ruta: ")
        for i in range(len(results[o-1])):
            if ( instance.Nod_Ref[results[o-1][i][1]] == 0 ):
                print("Del Rack ", instance.Nod_Ref[results[o-1][i][0]], " al rack ", instance.Nod_Ref[results[o-1][i][1]], "......Entrega embalador")
            else:
                print("Del Rack ", instance.Nod_Ref[results[o-1][i][0]], " al rack ", instance.Nod_Ref[results[o-1][i][1]], "......Referencia: ", results[o-1][i][1])
            #fi
        #rof
        print("\n")
        print("Ruta continua(Racks): ", cont_route[o-1])
        print("\n")
    #rof
    
#fed
    
def save_routes_Picking( instance ):
    
    nueva_ruta = {"id":0, "distanciaRuta": 0, "recorrido": [], "referencias": []}
    Rutas["distanciaTotal"] = pyo.value(instance.FO)
    
    for o in instance.O:
        nr = copy.deepcopy( nueva_ruta )
        nr["id"] = o
        nr["distanciaRuta"] = instance.Dist_ORD[o].value
        for i in instance.ORDENES[o]:
            for j in instance.ORDENES[o]:
                if ( instance.x[o,i,j].value == 1 ):
                    r = [i, j]
                    nr["recorrido"].append(r)
                #fi
            #rof
        #rof
        Rutas["ruta"].append(nr)
    #rof
    
    ## Ordenar Recorridos
    for i in range(len(Rutas["ruta"])):
        result = [Rutas["ruta"][i]["recorrido"][0]]
        ref = [Rutas["ruta"][i]["recorrido"][0][1]]
        for _ in range(len(Rutas["ruta"][i]["recorrido"])-1):
            aux = result[len(result)-1]
            for r in Rutas["ruta"][i]["recorrido"]:
                if ( r[0] == aux[1] ):
                    result.append(r)
                    ref.append(r[1])
                    break
                #fi
            #rof
        #rof
        Rutas["ruta"][i]["recorrido"] = result
        Rutas["ruta"][i]["referencias"] = ref
    #rof
    
    for i in range(len(Rutas["ruta"])):
        for x in Rutas["ruta"][i]["recorrido"]:
            x[0] = instance.Nod_Ref[x[0]]
            x[1] = instance.Nod_Ref[x[1]]
        #rof
    #rof
    print(Rutas)
    
    
#fed
################################################################################################################   
    
    
    
######################################################################################################################################  
def read_data_XLSX_Picking( receivedFile ):
    
    ## --------------------- READ DATA EXCEL WHIT OPENPYXL -----------------------------------
    
    ## OPEN FILE AND SAVE IN VARIABLE (archivo)
    archivo = openpyxl.load_workbook(receivedFile, data_only = True)

    ## READ SHEETS OF FILE AND SELECT THE THE FIRST SHEET
    sheets = archivo.sheetnames
    sheetDatos = archivo[sheets[0]]

    ## NODES (set_Nodos)
    set_Nodos = []
    column = sheetDatos['A']
    for x in range(len(column)): 
        if (type(column[x].value) != str and column[x].value != None):
            set_Nodos.append(column[x].value)
        #fi
    #rof

    ## ORDERS (set_Ords)
    set_Ords = []
    column = sheetDatos['B']
    for x in range(len(column)): 
        if (type(column[x].value) != str and column[x].value != None):
            set_Ords.append(column[x].value)
        #fi
    #rof


    ## REFERENCES (set_Referencias)
    set_Referencias = []
    column = sheetDatos['C']
    for x in range(len(column)): 
        if (type(column[x].value) != str and column[x].value != None):
            set_Referencias.append(column[x].value)
        #fi
    #rof

    ## LOCATION OF EACH REFERENCE (param_NOD_REF)
    lect = []
    column = sheetDatos['F']
    for x in range(len(column)): 
        if (type(column[x].value) != str and column[x].value != None):
            lect.append(column[x].value)
        #fi
    #rof
    param_NOD_REF = dict(zip(set_Referencias, lect))


    ## Ordenes y Variable auxiliar (set_Ordenes y set_R)
    lect = []
    for i in range(len(set_Ords)):
        column = sheetDatos[get_column_letter( 8 + i )]
        orden = []
        for x in range(len(column)): 
            if (type(column[x].value) != str and column[x].value != None):
                orden.append(column[x].value)
            #fi
        #rof
        lect.append(orden)
    #rof
    set_Ordenes = dict(zip(set_Ords, lect))

    lectR = copy.deepcopy(lect) ## Esto es demencial
    for x in lectR:
        x.remove(0)
    #rof
    set_R = dict(zip(set_Ords, lectR))


    ## DISTANCE MATRIX
    param_Distancia = []
    sheetDistancias = archivo[sheets[1]]
    for i in  range(sheetDistancias.min_row + 1,sheetDistancias.max_row + 1):
        y = []
        for j in range(sheetDistancias.min_column + 1, sheetDistancias.max_column + 1):
            y.append(sheetDistancias.cell(row = i, column = j).value)
        #rof
        param_Distancia.append(y)
    #rof
    ## -------------------------------------------------------------------------------
    
    return( set_Nodos, set_Ords, set_Referencias, param_NOD_REF, set_Ordenes, set_R, param_Distancia )
#fed
######################################################################################################################################    
    

######################################################################################################################################
def create_lineal_model_Picking( model, set_Nodos, set_Ords, set_Referencias, param_NOD_REF, set_Ordenes, set_R, param_Distancia ):
    

    ## ------------------- SETS --------------------------------
    model.NODOS = pyo.Set( initialize = set_Nodos )
    model.O = pyo.Set( initialize = set_Ords )
    model.REF = pyo.Set( initialize = set_Referencias )
    def setORDENES(model, i):
        return(list(set_Ordenes[i]))
    #fed
    model.ORDENES = pyo.Set(model.O, initialize = setORDENES)
    def setR(model, i):
        return(list(set_R[i]))
    #fed
    model.R = pyo.Set(model.O, initialize = set_R)



    def junte(model):
        return((o,i,j) for o in model.O for i in model.ORDENES[o] for j in model.ORDENES[o] )
    #fed
    model.OROR = pyo.Set(dimen =3, initialize = junte )

    def junte2(model):
        return((o,i) for o in model.O for i in model.ORDENES[o] )
    #fed
    model.OX = pyo.Set(dimen = 2, initialize = junte2 )

    def junte3(model):
        return((o,i,j) for o in model.O for i in model.R[o] for j in model.R[o] )
    #fed
    model.ORR = pyo.Set(dimen = 3, initialize = junte3 )

    def junte4(model):
        return((o,i) for o in model.O for i in model.R[o] )
    #fed
    model.OR  = pyo.Set(dimen = 2, initialize = junte4 )


    ## --------------------- PARAMETERS -------------------------------------
    def paramDistancia(model, i, j):
        return(param_Distancia[i][j])
    #fed
    model.Distancia = pyo.Param(model.NODOS,model.NODOS, initialize = paramDistancia)
    
    def paramNodRef(model, i):
        return(param_NOD_REF[i])
    #fed
    model.Nod_Ref = pyo.Param(model.REF, initialize = paramNodRef) 

    def paramDistR(model, i ,j):
        return model.Distancia[model.Nod_Ref[i],model.Nod_Ref[j]]
    #fed
    model.Distancia_R = pyo.Param(model.REF, model.REF, initialize = paramDistR, mutable = True)

    ## ----------------------- VARIABLES ---------------------------------------
    model.x = pyo.Var(model.OROR, domain = pyo.Binary)
    model.aux = pyo.Var(model.OR)
    model.Dist_ORD = pyo.Var(model.O)


    ## -------------------------- OBJECTIVE FUNCTION ------------------------------------
    def ObjFunc(model):
        return sum(model.Distancia_R[i,j]*model.x[o,i,j] for o in model.O for i in model.ORDENES[o] for j in model.ORDENES[o])
    #fed
    model.FO = pyo.Objective(rule = ObjFunc)


    ## --------------------------- RESTRICTIONS ----------------------------------------------
    def r1(model, o, i):
        return sum( model.x[o,i,j] for j in model.ORDENES[o]) == 1
    #fed
    model.r1 = pyo.Constraint( model.OX, rule = r1 )

    def r2(model, o, j):
        return sum( model.x[o,i,j] for i in model.ORDENES[o]) == 1
    #fed
    model.r2 = pyo.Constraint( model.OX,  rule = r2 )

    def r3(model, o, i):
        return model.x[o,i,i] == 0
    #fed
    model.r3 = pyo.Constraint( model.OX,  rule = r3 )

    def r4(model, o, i, j):
        if ( i != j ):
            return model.aux[o,i] - model.aux[o,j] + len(model.R[o])*model.x[o,i,j] <= len(model.R[o]) - 1 
        return pyo.Constraint.Skip
    #fed
    model.r4 = pyo.Constraint( model.ORR,  rule = r4 )

    def r5(model, o):
        return model.Dist_ORD[o] == sum( model.x[o,i,j]*model.Distancia_R[i,j] for i in model.ORDENES[o] for j in model.ORDENES[o] )
    #fed
    model.r5 = pyo.Constraint( model.O, rule = r5)
#fed
#########################################################################################################################################
    
    
    
    
    
#########################################################################################################################################
def print_results_XLSX_Picking( workbook ):
    ## -------------------------- SAVE RESULTS IN EXCEL FILE -------------------------------------

    ## STYLES
    merge_format = workbook.add_format({
        'bold': 1,
        'align': 'center',
        'valign': 'vcenter'})
    cell_format = workbook.add_format({'align': 'center'})
        
    ## CREATE SHEET
    worksheet = workbook.add_worksheet('Resultados')
    worksheet.set_column(0, 3, 11) ## Width: set_column(first_col, last_col, width, cell_format, options)
    worksheet.set_column(4, 4, 3)
    worksheet.set_column(5, 5, 20)
    
    ## PRINT RESULTS
    worksheet.merge_range(0, 0, 0, 3, "SOLUCIÓN DEL EJERCICIO", merge_format)
    worksheet.merge_range(1, 0, 1, 1, "Distancia Total: ", merge_format)
    worksheet.merge_range(1, 2, 1, 3, Rutas["distanciaTotal"], cell_format)
    
    row=3
    col=0
    for ruta in Rutas["ruta"]:
        worksheet.merge_range(row, col, row, 3, "Ruta "+ str(ruta["id"]), merge_format)
        row += 1
        
        worksheet.merge_range(row, col, row, col + 1,"Distancia Recorrida: ", merge_format)
        worksheet.merge_range(row, col + 2, row, col + 3,ruta["distanciaRuta"], cell_format)
        row += 1
        
        worksheet.merge_range(row, col, row, col + 3, "Recorrido: ", cell_format)
        worksheet.write(row, col + 5, "Seleccionar Ref: ", merge_format)
        row += 1
        
        for k in range(len(ruta["recorrido"])):
            if (ruta["referencias"][k] != 0):
                worksheet.write(row, col + 0,"Del rack ", cell_format)
                worksheet.write(row, col + 1, ruta["recorrido"][k][0], cell_format)
                worksheet.write(row, col + 2," al rack", cell_format)
                worksheet.write(row, col + 3, ruta["recorrido"][k][1], cell_format)
                worksheet.write(row, col + 5, ruta["referencias"][k], merge_format)
            else:
                worksheet.write(row, col + 0,"Del rack ", cell_format)
                worksheet.write(row, col + 1, ruta["recorrido"][k][0], cell_format)
                worksheet.write(row, col + 2," al rack", cell_format)
                worksheet.write(row, col + 3, ruta["recorrido"][k][1], cell_format)
                worksheet.write(row, col + 5, "Entregar al embalador", merge_format)
            #fi
            row += 1
        #rof
        row += 1
    #rof
        
        
    ## CLOSE THE BOOK
    workbook.close()
#fed
##########################################################################################################################################